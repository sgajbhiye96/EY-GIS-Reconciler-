
# # ==========================================================
# # EY GIS RECONCILER — FINAL VERSION (GIS → ORG CHART FUZZY)
# # ==========================================================

# import streamlit as st
# st.set_page_config(page_title="EY GIS Reconciler — Final", layout="wide")

# import pandas as pd
# import json
# import base64
# import tempfile
# import io
# import re
# from rapidfuzz import fuzz
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill, Font, Border, Side

# # Azure Vision
# try:
#     from openai import AzureOpenAI
# except:
#     AzureOpenAI = None

# try:
#     import fitz  # PyMuPDF
# except:
#     fitz = None

# # OCR fallback
# try:
#     from PIL import Image
#     import pytesseract
# except:
#     Image = None
#     pytesseract = None


# # -------------------------
# # SIDEBAR
# # -------------------------
# st.sidebar.header("Controls")
# FUZZY_THRESHOLD = st.sidebar.slider("Fuzzy threshold", 60, 100, 80, 1)


# # -------------------------
# # AZURE CONFIG
# # -------------------------
# AZURE_ENDPOINT = st.secrets.get("AZURE_OPENAI_ENDPOINT", None)
# AZURE_KEY = st.secrets.get("AZURE_OPENAI_KEY", None)
# AZURE_DEPLOY = st.secrets.get("AZURE_OPENAI_DEPLOYMENT", "gpt4o")
# AZURE_API_VER = st.secrets.get("AZURE_OPENAI_API_VERSION", "2025-03-01-preview")

# client = None
# if AZURE_KEY and AzureOpenAI:
#     client = AzureOpenAI(api_key=AZURE_KEY,
#                          api_version=AZURE_API_VER,
#                          azure_endpoint=AZURE_ENDPOINT)


# # -------------------------
# # NORMALIZATION
# # -------------------------
# def normalize_gis(raw):
#     """Use only columns Entity Name + Parent Name."""
#     cols = [c.lower().strip() for c in raw.columns]

#     if "entity name" in cols and "parent name" in cols:
#         ent = raw.columns[cols.index("entity name")]
#         par = raw.columns[cols.index("parent name")]
#         df = pd.DataFrame({
#             "entity": raw[ent].astype(str).str.strip(),
#             "parent": raw[par].astype(str).str.strip()
#         })
#         return df

#     st.error("GIS file must contain 'Entity Name' and 'Parent Name' columns.")
#     return pd.DataFrame(columns=["entity", "parent"])


# def normalize_client(df):
#     """Convert extraction JSON/CSV to entity + parent."""
#     if df is None or df.empty:
#         return pd.DataFrame(columns=["entity", "parent"])

#     cols = [c.lower() for c in df.columns]

#     ent_col = df.columns[0]
#     par_col = df.columns[1] if len(df.columns) > 1 else None

#     df_out = pd.DataFrame({
#         "entity": df[ent_col].astype(str).str.strip(),
#         "parent": df[par_col].astype(str).str.strip() if par_col else ""
#     })

#     return df_out


# # -------------------------
# # PDF → Images
# # -------------------------
# def pdf_to_images(pdf_bytes, dpi=250):
#     if fitz is None:
#         st.error("PyMuPDF not installed.")
#         return []
#     with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
#         tmp.write(pdf_bytes)
#         path = tmp.name
#     pdf = fitz.open(path)

#     images = []
#     for p in pdf:
#         pix = p.get_pixmap(dpi=dpi)
#         images.append(pix.tobytes("png"))
#     return images


# # -------------------------
# # Azure Vision Extractor
# # -------------------------
# def call_gpt4o_extract(image_bytes):
#     if client is None:
#         return None

#     img64 = base64.b64encode(image_bytes).decode()

#     try:
#         resp = client.responses.create(
#             model=AZURE_DEPLOY,
#             input=[{
#                 "role": "user",
#                 "content": [
#                     {"type": "input_text",
#                      "text": "Extract ALL entities and parents. "
#                              "Return STRICT JSON array: "
#                              "[{\"entity\":\"...\",\"parent\":\"...\"}, ...]"},
#                     {"type": "input_image",
#                      "image_url": f"data:image/png;base64,{img64}"}
#                 ]
#             }],
#             max_output_tokens=1400
#         )
#         text_out = []
#         for msg in resp.output:
#             for c in msg.content:
#                 if hasattr(c, "text"):
#                     text_out.append(c.text)
#         return "\n".join(text_out)
#     except Exception as e:
#         st.warning(f"Vision error: {e}")
#         return None


# # -------------------------
# # Parse JSON from model
# # -------------------------
# def parse_model_json(raw_text):
#     if not raw_text:
#         return None

#     cleaned = raw_text.replace("```json", "").replace("```", "").strip()

#     try:
#         obj = json.loads(cleaned)
#         df = pd.DataFrame(obj)
#         return normalize_client(df)
#     except:
#         pass

#     try:
#         obj = eval(cleaned.replace(": null", ": None"))
#         df = pd.DataFrame(obj)
#         return normalize_client(df)
#     except:
#         pass

#     return None


# # -------------------------
# # *** FINAL RECONCILIATION LOGIC ***
# # GIS → Org Chart only (as requested)
# # -------------------------
# def build_reconciliation(df_client, df_gis, fuzzy_threshold=80):
#     """
#     NEW VERSION:
#     - Fuzzy Matching is now applied ONLY on ORG CHART (df_client)
#     - GIS fuzzy matching removed completely
#     - Output contains fuzzy match entity + fuzzy match parent from ORG CHART
#     """

#     df_c = df_client.copy()
#     df_g = df_gis.copy()

#     df_c["entity_l"] = df_c["entity"].astype(str).str.lower().str.strip()
#     df_c["parent_l"] = df_c["parent"].astype(str).str.lower().str.strip()
#     df_g["entity_l"] = df_g["entity"].astype(str).str.lower().str.strip()
#     df_g["parent_l"] = df_g["parent"].astype(str).str.lower().str.strip()

#     all_entities = sorted(set(df_c["entity_l"]).union(set(df_g["entity_l"])))

#     rows = []
#     for ent in all_entities:
#         crows = df_c[df_c["entity_l"] == ent]
#         grows = df_g[df_g["entity_l"] == ent]

#         # extract names from both sides
#         client_name = crows.iloc[0]["entity"] if not crows.empty else ""
#         client_parent = crows.iloc[0]["parent"] if not crows.empty else ""

#         gis_name = grows.iloc[0]["entity"] if not grows.empty else ""
#         gis_parent = grows.iloc[0]["parent"] if not grows.empty else ""

#         # ----------------------------
#         # NEW → fuzzy match AGAINST ORG CHART ONLY
#         # ----------------------------
#         best_ent_match = ""
#         best_ent_score = 0

#         best_par_match = ""
#         best_par_score = 0

#         for _, r in df_c.iterrows():

#             # entity name fuzzy
#             e_score = fuzz.token_sort_ratio(str(gis_name).lower(), str(r["entity"]).lower())
#             if e_score > best_ent_score:
#                 best_ent_score = e_score
#                 best_ent_match = r["entity"]

#             # parent fuzzy
#             p_score = fuzz.token_sort_ratio(str(gis_parent).lower(), str(r["parent"]).lower())
#             if p_score > best_par_score:
#                 best_par_score = p_score
#                 best_par_match = r["parent"]

#         # ----------------------------
#         # ACTION LOGIC (unchanged)
#         # ----------------------------
#         in_client = bool(client_name.strip())
#         in_gis = bool(gis_name.strip())

#         if in_client and not in_gis:
#             action = "To be added in GIS."
#             color = "yellow"
#         elif in_gis and not in_client:
#             action = "To be removed from GIS."
#             color = "red"
#         else:
#             if str(client_parent).lower().strip() == str(gis_parent).lower().strip():
#                 action = "Matched"
#                 color = "white"
#             else:
#                 action = "Mismatch noted, please further check."
#                 color = "red"

#         rows.append({
#             "Entity Name in GIS": gis_name,
#             "Entity Name in Org Chart": client_name,
#             "Parent Name in GIS": gis_parent,
#             "Parent Name in Org Chart": client_parent,

#             # NEW fuzzy columns using ORG CHART only
#             "Fuzzy Match (Org Entity)": best_ent_match,
#             "Fuzzy Match (Org Parent)": best_par_match,
#             "Fuzzy Score (Entity)": best_ent_score,
#             "Fuzzy Score (Parent)": best_par_score,

#             "Action point": action,
#             "_color": color
#         })

#     return pd.DataFrame(rows)

# # -------------------------
# # Excel Export
# # -------------------------
# def export_excel(recon_df):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Reconciliation"

#     headers = [
#         "Entity Name in GIS",
#         "Entity Name in Org Chart",
#         "Parent Name in GIS",
#         "Parent Name in Org Chart",
#         "Fuzzy Match (Org Entity)",
#         "Fuzzy Match (Org Parent)",
#         "Fuzzy Score (Entity)",
#         "Fuzzy Score (Parent)",
#         "Action point"
#     ]
#     ws.append(headers)

#     yellow = "FFF2CC"
#     red = "F8CBCC"
#     white = "FFFFFF"
#     border = Border(left=Side(style="thin"), right=Side(style="thin"),
#                     top=Side(style="thin"), bottom=Side(style="thin"))

#     for _, r in recon_df.iterrows():
#         ws.append([
#             r.get("Entity Name in GIS", ""),
#             r.get("Entity Name in Org Chart", ""),
#             r.get("Parent Name in GIS", ""),
#             r.get("Parent Name in Org Chart", ""),
#             r.get("Fuzzy Match (Org Entity)", ""),
#             r.get("Fuzzy Match (Org Parent)", ""),
#             r.get("Fuzzy Score (Entity)", ""),
#             r.get("Fuzzy Score (Parent)", ""),
#             r.get("Action point", "")
#         ])

#     # header style
#     for cell in ws[1]:
#         cell.fill = PatternFill(start_color="FFD700", fill_type="solid")
#         cell.font = Font(bold=True)
#         cell.border = border

#     # row coloring
#     for i in range(2, len(recon_df) + 2):
#         color = recon_df.iloc[i-2].get("_color", "white")
#         fill_color = white if color == "white" else red if color == "red" else yellow
#         for cell in ws[i]:
#             cell.fill = PatternFill(start_color=fill_color, fill_type="solid")
#             cell.border = border

#     # save
#     path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
#     wb.save(path)
#     return path

# # -------------------------
# # UI
# # -------------------------
# st.title("EY GIS Reconciler — GIS → Org Chart (Fuzzy Only Within Org Chart)")

# client_file = st.file_uploader("Upload Org Chart (PNG/JPG/PDF/CSV/JSON)", type=["png","jpg","jpeg","pdf","csv","json"])
# gis_file = st.file_uploader("Upload GIS Extract", type=["xlsx","csv"])

# if not gis_file:
#     st.stop()

# # Load GIS
# if gis_file.name.endswith(".xlsx"):
#     raw_gis = pd.read_excel(gis_file)
# else:
#     raw_gis = pd.read_csv(gis_file)

# df_gis = normalize_gis(raw_gis)


# # -------------------------
# # Extract client org chart
# # -------------------------
# df_client = pd.DataFrame(columns=["entity", "parent"])
# pages = []

# if client_file:
#     data = client_file.read()

#     # PDF
#     if client_file.name.endswith(".pdf"):
#         images = pdf_to_images(data, dpi=250)
#         for img in images:
#             raw = call_gpt4o_extract(img)
#             df = parse_model_json(raw)
#             if df is not None:
#                 pages.append(df)

#     # Images
#     elif client_file.name.lower().endswith(("png","jpg","jpeg")):
#         raw = call_gpt4o_extract(data)
#         df = parse_model_json(raw)
#         if df is not None:
#             pages.append(df)

#     # CSV / JSON
#     else:
#         try:
#             if client_file.name.endswith(".json"):
#                 df_client_raw = pd.DataFrame(json.loads(data))
#             else:
#                 df_client_raw = pd.read_csv(io.BytesIO(data))
#             pages.append(normalize_client(df_client_raw))
#         except:
#             pass

# if pages:
#     df_client = pd.concat(pages, ignore_index=True)


# # Show preview
# st.subheader("Extracted Org Chart (Client)")
# st.dataframe(df_client)


# # -------------------------
# # RUN RECONCILIATION
# # -------------------------
# st.subheader("Reconciliation (GIS → Org Chart)")
# recon = build_reconciliation(df_client, df_gis, fuzzy_threshold=FUZZY_THRESHOLD)

# st.dataframe(recon.drop(columns=["_color"]), height=500)


# # -------------------------
# # EXPORT BUTTON
# # -------------------------
# path = export_excel(recon)
# with open(path, "rb") as f:
#     st.download_button("Download Reconciliation Excel", f, "reconciliation.xlsx")

# app.py
import streamlit as st
st.set_page_config(page_title="EY GIS Reconciler — Org→GIS Fuzzy (Cleaned)", layout="wide")

import pandas as pd
import json
import base64
import tempfile
import io
import re
import unicodedata
from rapidfuzz import fuzz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side

# Optional Azure Responses client
try:
    from openai import AzureOpenAI
except Exception:
    AzureOpenAI = None

# PDF handling
try:
    import fitz  # pymupdf
except Exception:
    fitz = None

# OCR fallback
try:
    from PIL import Image
    import pytesseract
except Exception:
    Image = None
    pytesseract = None

# ---------------------------
# Sidebar controls
# ---------------------------
st.sidebar.header("Controls")
FUZZY_DEFAULT = 80
FUZZY_THRESHOLD = st.sidebar.slider("Fuzzy threshold (entity)", 60, 100, FUZZY_DEFAULT, 1)
PARENT_FUZZY_THRESHOLD = st.sidebar.slider("Fuzzy threshold (parent fallback)", 60, 100, 85, 1)
st.sidebar.markdown("Fuzzy = Org Chart → GIS (Option A). Parent matching uses exact match first, fuzzy fallback.")

# ---------------------------
# Azure/OpenAI config (optional)
# ---------------------------
AZURE_ENDPOINT = st.secrets.get("AZURE_OPENAI_ENDPOINT", None)
AZURE_KEY = st.secrets.get("AZURE_OPENAI_KEY", None)
AZURE_DEPLOY = st.secrets.get("AZURE_OPENAI_DEPLOYMENT", "gpt4o")
AZURE_API_VER = st.secrets.get("AZURE_OPENAI_API_VERSION", "2025-03-01-preview")

client = None
if AZURE_KEY and AzureOpenAI is not None and AZURE_ENDPOINT:
    client = AzureOpenAI(api_key=AZURE_KEY, api_version=AZURE_API_VER, azure_endpoint=AZURE_ENDPOINT)

# ---------------------------
# Utility: ultra_clean
# ---------------------------
def ultra_clean(x):
    """Normalize string aggressively to reduce OCR/unicode issues while preserving visible text (keeps parentheses)."""
    if pd.isna(x):
        return ""
    s = str(x)
    # Unicode normalization
    s = unicodedata.normalize("NFKC", s)
    # Remove zero-width and BOM
    s = s.replace("\u200b", "").replace("\u200c", "").replace("\u200d", "").replace("\ufeff", "")
    # Replace non-breaking spaces with normal spaces
    s = s.replace("\u00A0", " ")
    # Replace various dash types with standard hyphen
    s = re.sub(r"[–—−]", "-", s)
    # Collapse multiple whitespace/newlines
    s = re.sub(r"\s+", " ", s)
    # Trim
    s = s.strip()
    return s

# ---------------------------
# Normalizers for incoming files
# ---------------------------
def normalize_gis_dataframe(raw):
    """Map GIS file to DataFrame with columns: entity, parent (keeps parentheses). Applies ultra_clean."""
    if raw is None or raw.empty:
        return pd.DataFrame(columns=["entity", "parent"])
    cols = [str(c).strip() for c in raw.columns]
    lower = [c.lower() for c in cols]
    if "entity name" in lower and "parent name" in lower:
        ent_col = cols[lower.index("entity name")]
        par_col = cols[lower.index("parent name")]
    else:
        # fallback: try common names
        ent_col = None
        par_col = None
        for k in ["entity", "name", "company"]:
            if k in lower:
                ent_col = cols[lower.index(k)]
                break
        for k in ["parent", "parent name", "parent_name"]:
            if k in lower:
                par_col = cols[lower.index(k)]
                break
        if ent_col is None:
            ent_col = cols[0]
        if par_col is None:
            par_col = cols[1] if len(cols) > 1 else None

    ent_series = raw[ent_col].astype(str).apply(ultra_clean)
    par_series = raw[par_col].astype(str).apply(ultra_clean) if par_col is not None else pd.Series([""] * len(ent_series))
    return pd.DataFrame({"entity": ent_series, "parent": par_series})

def normalize_client_df(df):
    """Normalize extracted client dataframe to columns entity,parent and apply ultra_clean."""
    if df is None or df.empty:
        return pd.DataFrame(columns=["entity", "parent"])
    cols = [str(c).strip() for c in df.columns]
    lower = [c.lower() for c in cols]

    ent_col = None
    par_col = None
    for k in ["entity", "name", "child", "node", "company"]:
        if k in lower:
            ent_col = cols[lower.index(k)]
            break
    for k in ["parent", "parent name", "parent_name"]:
        if k in lower:
            par_col = cols[lower.index(k)]
            break
    if ent_col is None:
        ent_col = cols[0]
    if par_col is None:
        par_col = cols[1] if len(cols) > 1 else None

    ent_series = df[ent_col].astype(str).apply(ultra_clean)
    par_series = df[par_col].astype(str).apply(ultra_clean) if par_col else pd.Series([""] * len(ent_series))
    return pd.DataFrame({"entity": ent_series, "parent": par_series})

# ---------------------------
# PDF -> images bytes (PyMuPDF)
# ---------------------------
def pdf_to_images_bytes(pdf_bytes, dpi=200):
    if fitz is None:
        raise RuntimeError("pymupdf (fitz) required for PDF processing. Install pymupdf.")
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name
    pdf = fitz.open(tmp_path)
    images = []
    for page in pdf:
        pix = page.get_pixmap(dpi=dpi)
        images.append(pix.tobytes("png"))
    return images

# ---------------------------
# Azure Vision call (Responses)
# ---------------------------
def call_gpt4o_extract(image_bytes, max_output_tokens=1500):
    """Call Azure Responses (vision) to extract JSON-like entity/parent array. Returns text output or None."""
    if client is None:
        return None
    img_b64 = base64.b64encode(image_bytes).decode()
    try:
        resp = client.responses.create(
            model=AZURE_DEPLOY,
            input=[{
                "role": "user",
                "content": [
                    {"type": "input_text", "text": (
                        "Extract ALL entities and parent-child relationships from this organisation chart image. "
                        "Return STRICT JSON array EXACTLY like: [{\"entity\":\"<entity name>\", \"parent\":\"<parent name>\"}, ...]. "
                        "Top-level entities should have parent as an empty string. Return ONLY the JSON array."
                    )},
                    {"type": "input_image", "image_url": f"data:image/png;base64,{img_b64}"}
                ]
            }],
            max_output_tokens=max_output_tokens
        )
        texts = []
        if resp.output:
            for msg in resp.output:
                for c in msg.content:
                    if hasattr(c, "text") and c.text:
                        texts.append(c.text)
        return "\n".join(texts) if texts else None
    except Exception as e:
        st.warning(f"Vision API error: {e}")
        return None

# ---------------------------
# Robust model output parser
# ---------------------------
def parse_model_json(raw_text):
    """Try to parse strict JSON from model response, fallback to heuristics. Returns normalized client df or None."""
    if not raw_text or not raw_text.strip():
        return None
    cleaned = raw_text.strip().replace("```json", "").replace("```", "").strip()
    # try json loads
    try:
        obj = json.loads(cleaned)
        if isinstance(obj, dict):
            # try to find a list value
            for v in obj.values():
                if isinstance(v, list):
                    obj = v
                    break
        if isinstance(obj, list):
            df = pd.DataFrame(obj)
            return normalize_client_df(df)
    except Exception:
        pass
    # try eval
    try:
        py = cleaned.replace(": null", ": None")
        obj = eval(py, {"__builtins__": {}})
        if isinstance(obj, list):
            df = pd.DataFrame(obj)
            return normalize_client_df(df)
    except Exception:
        pass
    # heuristic parse lines
    pairs = []
    lines = [ln.strip() for ln in cleaned.splitlines() if ln.strip()]
    i = 0
    while i < len(lines):
        ln = lines[i]
        m = re.match(r'^(?:entity|name)\s*[:\-]\s*(.+)$', ln, flags=re.I)
        if m:
            ent = m.group(1).strip()
            parent = ""
            if i+1 < len(lines):
                m2 = re.match(r'^(?:parent)\s*[:\-]\s*(.+)$', lines[i+1], flags=re.I)
                if m2:
                    parent = m2.group(1).strip()
                    i += 1
            pairs.append({"entity": ent, "parent": parent})
            i += 1
            continue
        if "," in ln:
            parts = [p.strip() for p in ln.split(",")]
            if len(parts) >= 2:
                pairs.append({"entity": parts[0], "parent": parts[1]})
                i += 1
                continue
        parts = re.split(r'\s+[—–-]\s+', ln)
        if len(parts) == 2:
            pairs.append({"entity": parts[0].strip(), "parent": parts[1].strip()})
            i += 1
            continue
        pairs.append({"entity": ln, "parent": ""})
        i += 1
    if pairs:
        df = pd.DataFrame(pairs)
        return normalize_client_df(df)
    return None

# ---------------------------
# Core reconciliation (Option A — Org Chart → GIS fuzzy)
# ---------------------------
def build_reconciliation(df_client, df_gis, fuzzy_threshold=80, parent_fuzzy_threshold=85):
    """
    For each Org Chart (client) entity:
      - Try exact match to GIS entity (cleaned)
      - If exact found: compare parent (exact -> Matched; else fallback to parent fuzzy)
      - If no exact GIS entity: fuzzy match GIS entity names and pick best (if >= fuzzy_threshold) else "To be removed from GIS tree." per your rule
    After processing client entities, add GIS-only entities (not present in client) as "To be added in GIS tree."
    """
    # Defensive copies
    df_c = df_client.copy().reset_index(drop=True)
    df_g = df_gis.copy().reset_index(drop=True)

    # Ensure cleaned columns exist
    df_c["entity_clean"] = df_c["entity"].apply(ultra_clean)
    df_c["parent_clean"] = df_c["parent"].apply(ultra_clean)
    df_g["entity_clean"] = df_g["entity"].apply(ultra_clean)
    df_g["parent_clean"] = df_g["parent"].apply(ultra_clean)

    df_c["entity_l"] = df_c["entity_clean"].str.lower()
    df_c["parent_l"] = df_c["parent_clean"].str.lower()
    df_g["entity_l"] = df_g["entity_clean"].str.lower()
    df_g["parent_l"] = df_g["parent_clean"].str.lower()

    rows = []
    # Keep track of which GIS rows have been matched (by index)
    matched_gis_idxs = set()

    # Process each Org Chart entity (client-driven matching)
    for ci, crow in df_c.iterrows():
        client_e = crow["entity"]
        client_p = crow["parent"]
        client_e_l = crow["entity_l"]
        client_p_l = crow["parent_l"]

        # 1) Exact match on GIS entity (cleaned)
        gis_exact = df_g[df_g["entity_l"] == client_e_l]
        if not gis_exact.empty:
            # pick first if multiple
            gi = gis_exact.index[0]
            g_row = df_g.loc[gi]
            matched_gis_idxs.add(gi)

            gis_e = g_row["entity"]
            gis_p = g_row["parent"]
            gis_p_l = g_row["parent_l"]

            # Parent exact check first
            if client_p_l == gis_p_l:
                action = "Matched"
                color = "white"
                fuzzy_ent = gis_e
                fuzzy_par = gis_p
                fs_ent = 100
                fs_par = 100
            else:
                # Parent mismatch -> try parent fuzzy fallback (compare client parent vs GIS parent list)
                # We'll compute a fuzzy score between client parent and the GIS parent we matched
                par_score = fuzz.token_sort_ratio(client_p_l, gis_p_l) if client_p_l and gis_p_l else 0
                if par_score >= parent_fuzzy_threshold:
                    action = "Matched (Parent fuzzy fallback)"
                    color = "white"
                else:
                    action = "Mismatch noted, please further check."
                    color = "red"
                fuzzy_ent = gis_e
                fuzzy_par = gis_p
                fs_ent = 100
                fs_par = par_score

            rows.append({
                "Entity Name in Org Chart": client_e,
                "Entity Name in GIS (Exact)": gis_e,
                "Parent Name in Org Chart": client_p,
                "Parent Name in GIS (Exact)": gis_p,
                "Fuzzy Best Match (Org→GIS)": fuzzy_ent,
                "Fuzzy Best Match Parent (Org→GIS)": fuzzy_par,
                "Fuzzy Score (Entity)": int(fs_ent),
                "Fuzzy Score (Parent)": int(fs_par),
                "Action point": action,
                "_color": color
            })
            continue

        # 2) No exact GIS entity -> fuzzy match across GIS entity names (Org Chart -> GIS)
        best_score = 0
        best_idx = None
        for gi, g in df_g.iterrows():
            score = fuzz.token_sort_ratio(client_e_l, g["entity_l"])
            if score > best_score:
                best_score = score
                best_idx = gi

        if best_score >= fuzzy_threshold and best_idx is not None:
            # Accept fuzzy matched GIS entity as candidate
            g_row = df_g.loc[best_idx]
            matched_gis_idxs.add(best_idx)

            gis_e = g_row["entity"]
            gis_p = g_row["parent"]
            gis_p_l = g_row["parent_l"]

            # For parent: first exact check vs matched GIS parent; if fails, compute parent fuzzy
            if client_p_l == gis_p_l:
                action = "Matched (Fuzzy on entity)"
                color = "white"
            else:
                par_score = fuzz.token_sort_ratio(client_p_l, gis_p_l) if client_p_l and gis_p_l else 0
                if par_score >= parent_fuzzy_threshold:
                    action = "Matched (Entity fuzzy + Parent fuzzy fallback)"
                    color = "white"
                else:
                    action = "Mismatch noted, please further check."
                    color = "red"

            rows.append({
                "Entity Name in Org Chart": client_e,
                "Entity Name in GIS (Exact)": gis_e,
                "Parent Name in Org Chart": client_p,
                "Parent Name in GIS (Exact)": gis_p,
                "Fuzzy Best Match (Org→GIS)": gis_e,
                "Fuzzy Best Match Parent (Org→GIS)": gis_p,
                "Fuzzy Score (Entity)": int(best_score),
                "Fuzzy Score (Parent)": int(par_score if 'par_score' in locals() else 0),
                "Action point": action,
                "_color": color
            })
        else:
            # Not found in GIS (client exists but GIS doesn't): per your specified rule => "To be removed from GIS tree."
            rows.append({
                "Entity Name in Org Chart": client_e,
                "Entity Name in GIS (Exact)": "",
                "Parent Name in Org Chart": client_p,
                "Parent Name in GIS (Exact)": "",
                "Fuzzy Best Match (Org→GIS)": "",
                "Fuzzy Best Match Parent (Org→GIS)": "",
                "Fuzzy Score (Entity)": int(best_score),
                "Fuzzy Score (Parent)": 0,
                "Action point": "To be removed from GIS tree.",
                "_color": "yellow"
            })

    # 3) GIS-only entities -> To be added in GIS tree.
    for gi, g in df_g.iterrows():
        if gi in matched_gis_idxs:
            continue
        rows.append({
            "Entity Name in Org Chart": "",
            "Entity Name in GIS (Exact)": g["entity"],
            "Parent Name in Org Chart": "",
            "Parent Name in GIS (Exact)": g["parent"],
            "Fuzzy Best Match (Org→GIS)": "",
            "Fuzzy Best Match Parent (Org→GIS)": "",
            "Fuzzy Score (Entity)": 0,
            "Fuzzy Score (Parent)": 0,
            "Action point": "To be added in GIS tree.",
            "_color": "red"
        })

    recon = pd.DataFrame(rows)
    # Ensure all columns exist
    expected = [
        "Entity Name in Org Chart", "Entity Name in GIS (Exact)",
        "Parent Name in Org Chart", "Parent Name in GIS (Exact)",
        "Fuzzy Best Match (Org→GIS)", "Fuzzy Best Match Parent (Org→GIS)",
        "Fuzzy Score (Entity)", "Fuzzy Score (Parent)",
        "Action point", "_color"
    ]
    for c in expected:
        if c not in recon.columns:
            recon[c] = ""
    return recon[expected]

# ---------------------------
# Export styled Excel
# ---------------------------
def export_styled_excel(recon_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    headers = [
        "Entity Name in Org Chart", "Entity Name in GIS (Exact)",
        "Parent Name in Org Chart", "Parent Name in GIS (Exact)",
        "Fuzzy Best Match (Org→GIS)", "Fuzzy Best Match Parent (Org→GIS)",
        "Fuzzy Score (Entity)", "Fuzzy Score (Parent)", "Action point"
    ]
    ws.append(headers)

    yellow = "FFF2CC"
    red = "F8CBCC"
    white = "FFFFFF"
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for _, r in recon_df.iterrows():
        ws.append([
            r.get("Entity Name in Org Chart", ""),
            r.get("Entity Name in GIS (Exact)", ""),
            r.get("Parent Name in Org Chart", ""),
            r.get("Parent Name in GIS (Exact)", ""),
            r.get("Fuzzy Best Match (Org→GIS)", ""),
            r.get("Fuzzy Best Match Parent (Org→GIS)", ""),
            r.get("Fuzzy Score (Entity)", ""),
            r.get("Fuzzy Score (Parent)", ""),
            r.get("Action point", "")
        ])

    # style header
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="FFD700", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = border

    # style rows
    for i in range(2, len(recon_df) + 2):
        color = recon_df.iloc[i-2].get("_color", "white")
        fill_color = white
        if color == "yellow":
            fill_color = yellow
        elif color == "red":
            fill_color = red
        for cell in ws[i]:
            cell.fill = PatternFill(start_color=fill_color, fill_type="solid")
            cell.border = border

    path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
    wb.save(path)
    return path

# ---------------------------
# Streamlit UI
# ---------------------------
st.title("📊 EY GIS Reconciler — Org Chart → GIS Fuzzy (Option A)")

st.markdown("""
**Instructions**
- Upload the **GIS extract (Excel or CSV)** — it should contain `Entity Name` and `Parent Name` (case-insensitive).
- Upload the **Client Org Chart** as PNG/JPG/PDF or a pre-extracted CSV/JSON containing `entity,parent`.
- This tool normalizes names aggressively to improve matching (keeps parentheses).
""")

client_file = st.file_uploader("Client org chart (PNG/JPG/PDF/CSV/JSON)", type=["png","jpg","jpeg","pdf","csv","json"])
gis_file = st.file_uploader("GIS Extract (Excel/CSV) — must contain 'Entity Name' & 'Parent Name'", type=["xlsx","csv"])

if not gis_file:
    st.stop()

# Load GIS
try:
    if gis_file.name.lower().endswith(".xlsx"):
        raw_gis = pd.read_excel(gis_file)
    else:
        raw_gis = pd.read_csv(gis_file)
except Exception as e:
    st.error(f"Failed to read GIS file: {e}")
    st.stop()

df_gis = normalize_gis_dataframe(raw_gis)

# Load/extract client entities
pages = []
raw_outputs = []

if client_file:
    data = client_file.read()
    name = client_file.name.lower()
    try:
        if name.endswith(".pdf"):
            if fitz is None:
                st.error("pymupdf not installed — cannot process PDF. Install pymupdf.")
            else:
                images = pdf_to_images_bytes(data, dpi=250)
                for i, img in enumerate(images, start=1):
                    st.write(f"Processing page {i} ...")
                    raw_text = None
                    if client is not None:
                        raw_text = call_gpt4o_extract(img)
                    if raw_text:
                        raw_outputs.append(raw_text)
                        parsed = parse_model_json(raw_text)
                        if parsed is not None:
                            pages.append(parsed)
                    else:
                        # OCR fallback if available
                        if Image and pytesseract:
                            pil = Image.open(io.BytesIO(img))
                            ocr_text = pytesseract.image_to_string(pil)
                            if ocr_text:
                                raw_outputs.append(ocr_text)
                                parsed = parse_model_json(ocr_text)
                                if parsed is not None:
                                    pages.append(parsed)
        elif name.endswith((".png","jpg","jpeg")):
            raw_text = None
            if client is not None:
                raw_text = call_gpt4o_extract(data)
            if raw_text:
                raw_outputs.append(raw_text)
                parsed = parse_model_json(raw_text)
                if parsed is not None:
                    pages.append(parsed)
            else:
                if Image and pytesseract:
                    pil = Image.open(io.BytesIO(data))
                    ocr_text = pytesseract.image_to_string(pil)
                    if ocr_text:
                        raw_outputs.append(ocr_text)
                        parsed = parse_model_json(ocr_text)
                        if parsed is not None:
                            pages.append(parsed)
        else:
            # csv / json
            if name.endswith(".json"):
                obj = json.loads(data.decode("utf-8"))
                df_client_raw = pd.DataFrame(obj)
            else:
                df_client_raw = pd.read_csv(io.BytesIO(data))
            pages.append(normalize_client_df(df_client_raw))
    except Exception as e:
        st.error(f"Failed to process client file: {e}")

if pages:
    df_client = pd.concat(pages, ignore_index=True)
    df_client = normalize_client_df(df_client)
else:
    df_client = pd.DataFrame(columns=["entity", "parent"])
    st.warning("No extracted client entities — proceeding with empty client tree.")

# show preview
st.subheader("Preview — extracted Org Chart (client)")
st.dataframe(df_client.head(500))

if raw_outputs:
    with st.expander("Raw Vision / OCR outputs"):
        for i, r in enumerate(raw_outputs, start=1):
            st.write(f"--- Page {i} ---")
            st.code(r[:3000])

# Persist session state client/gis
if "df_client" not in st.session_state:
    st.session_state.df_client = df_client.copy()
if "df_gis" not in st.session_state:
    st.session_state.df_gis = df_gis.copy()

# Reconciliation
st.subheader("Reconciliation")
recon = build_reconciliation(st.session_state.df_client, st.session_state.df_gis,
                             fuzzy_threshold=FUZZY_THRESHOLD, parent_fuzzy_threshold=PARENT_FUZZY_THRESHOLD)

display_cols = [
    "Entity Name in Org Chart", "Entity Name in GIS (Exact)",
    "Parent Name in Org Chart", "Parent Name in GIS (Exact)",
    "Fuzzy Best Match (Org→GIS)", "Fuzzy Best Match Parent (Org→GIS)",
    "Fuzzy Score (Entity)", "Fuzzy Score (Parent)", "Action point"
]
st.dataframe(recon[display_cols], height=600)

# Human approval for fuzzy suggestions: allow accepting a fuzzy GIS match for a client row
st.subheader("Human approve fuzzy matches (Org→GIS)")
# candidates: client rows where GIS exact is empty but fuzzy entity exists & meets threshold
candidates = recon[
    (recon["Entity Name in Org Chart"].str.strip() != "") &
    (recon["Entity Name in GIS (Exact)"].str.strip() == "") &
    (recon["Fuzzy Score (Entity)"].fillna(0).astype(int) >= FUZZY_THRESHOLD)
].reset_index()

if not candidates.empty:
    cand_display = candidates.apply(
        lambda r: f"{r['Entity Name in Org Chart']}  → fuzzy GIS: {r['Fuzzy Best Match (Org→GIS)']} (score {r['Fuzzy Score (Entity)']})",
        axis=1
    ).tolist()
    sel = st.selectbox("Choose candidate to accept (adds GIS match into client session tree)", options=cand_display)
    if st.button("Accept selected fuzzy mapping (add GIS match as client match)"):
        idx = cand_display.index(sel)
        row = candidates.iloc[idx]
        # Add matched GIS pair into session client (so subsequent reconciliation shows exact match)
        new_entity = row["Fuzzy Best Match (Org→GIS)"]
        new_parent = row["Fuzzy Best Match Parent (Org→GIS)"]
        # Append to df_client in session
        st.session_state.df_client = pd.concat(
            [st.session_state.df_client, pd.DataFrame([{"entity": new_entity, "parent": new_parent}])],
            ignore_index=True
        )
        st.success(f"Accepted fuzzy mapping: added '{new_entity}' with parent '{new_parent}' to client tree (in-session).")
        st.experimental_rerun()
else:
    st.info("No fuzzy candidates (Org→GIS) meeting threshold and missing exact GIS match.")

# Export styled Excel
st.subheader("Export reconciliation")
excel_path = export_styled_excel(recon)
with open(excel_path, "rb") as f:
    st.download_button("Download reconciliation (styled Excel)", f, file_name="reconciliation.xlsx")

# Export updated client CSV after approvals
st.subheader("Download updated client (after approvals)")
if not st.session_state.df_client.empty:
    tmp_csv = tempfile.NamedTemporaryFile(delete=False, suffix=".csv").name
    st.session_state.df_client.to_csv(tmp_csv, index=False)
    with open(tmp_csv, "rb") as f:
        st.download_button("Download updated client CSV", f, file_name="client_updated.csv")

st.write("Done — uses only Entity Name & Parent Name. Fuzzy matching: Org Chart → GIS (Option A). Parent logic: exact first, fuzzy fallback.")
